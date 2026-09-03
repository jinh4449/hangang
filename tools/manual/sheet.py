"""매뉴얼 마크다운을 구글 시트용 엑셀 파일로 만든다.

    python3 tools/manual/sheet.py

나온 .xlsx 를 구글 드라이브에 올리고 「Google 스프레드시트로 열기」 하면 시트가 된다.
손으로 옮기지 않고 docs/manual/ 의 마크다운에서 직접 뽑는다. 손으로 옮기면 빠뜨린다.
"""
"""매뉴얼 마크다운을 시트에 넣을 행으로 푼다. 손으로 옮기면 빠뜨리므로 원본에서 뽑는다."""
import re, os, json

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.abspath(os.path.join(HERE, '..', '..'))
SRC = os.path.join(ROOT, 'docs', 'manual')
DOCS = [
    ('데스크',        'desk.md'),
    ('치료실',        'treatment-room.md'),
    ('한약·탕전',     'herbal.md'),
    ('원무·정산·재고', 'admin.md'),
]

def parse(fn):
    lines = open(os.path.join(SRC, fn), encoding='utf-8').read().split('\n')
    rows, tables = [], []
    sec = sub = ''
    i = 0
    skip_toc = False
    while i < len(lines):
        raw = lines[i]
        s = raw.strip()

        if s == '## 목차':                       # 시트에서는 필터가 목차 노릇을 한다
            skip_toc = True; i += 1; continue
        if skip_toc:
            if s.startswith('## '): skip_toc = False
            else: i += 1; continue

        if s.startswith('## '):
            sec, sub = s[3:].strip(), ''; i += 1; continue
        if s.startswith('### '):
            sub = s[4:].strip(); i += 1; continue
        if s.startswith('# ') or s == '---' or s == '' or (s.startswith('[') and '](./' in s and '·' in s):
            i += 1; continue

        # 표
        if s.startswith('|'):
            block = []
            while i < len(lines) and lines[i].strip().startswith('|'):
                block.append([c.strip() for c in lines[i].strip().strip('|').split('|')]); i += 1
            head, body = block[0], block[2:]
            tables.append({'sec': sec, 'sub': sub, 'head': head, 'body': body})
            for r in body:
                pairs = ' · '.join(f'{h}: {v}' for h, v in zip(head, r) if v)
                rows.append((sec, sub, '표', pairs))
            continue

        # 그림
        m = re.fullmatch(r'!\[([^\]]*)\]\(([^)]+)\)', s)
        if m:
            rows.append((sec, sub, '그림', f'{m.group(1)} ({m.group(2).split("/")[-1]})')); i += 1; continue

        # 목록 — 이어지는 들여쓴 줄은 같은 항목이다
        m = re.match(r'^(?:- |(\d+)\.\s+)(.*)$', s)
        if m:
            num, text = m.group(1), m.group(2)
            parts = [text]
            i += 1
            while i < len(lines) and lines[i].startswith('  ') and lines[i].strip():
                parts.append(lines[i].strip()); i += 1
            body = '\n'.join(parts)
            kind = '순서' if num else '항목'
            if parts[0].startswith('멘트'): kind = '멘트'
            elif parts[0].startswith('※') or parts[0].startswith('TIP)'): kind = '주의'
            label = f'{num}. ' if num else ''
            rows.append((sec, sub, kind, label + body))
            continue

        # 문단 — 멘트·※·Q/A 는 줄 단위로 끊는다
        para = []
        while i < len(lines) and lines[i].strip() and not lines[i].strip().startswith(('|', '- ', '#', '---')) \
              and not re.match(r'^\d+\.\s', lines[i].strip()):
            para.append(lines[i].strip()); i += 1
        buf, kind = [], '안내'
        def flush():
            if buf: rows.append((sec, sub, kind, '\n'.join(buf)))
        for l in para:
            if l.startswith('멘트'):
                flush(); buf, kind = [re.sub(r'^멘트\s*', '', l)], '멘트'
            elif l.startswith('※') or l.startswith('TIP)'):
                flush(); buf, kind = [l], '주의'
            elif l.startswith('Q. '):
                flush(); buf, kind = [l], '문답'
            elif kind in ('멘트', '주의', '문답'):
                buf.append(l)
            else:
                if kind != '안내': flush(); buf, kind = [], '안내'
                buf.append(l)
        flush()
    return rows, tables

out = {'docs': {}, 'tables': []}
for name, fn in DOCS:
    rows, tables = parse(fn)
    out['docs'][name] = rows
    for t in tables:
        t['doc'] = name
        out['tables'].append(t)

D = out


import re, math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = os.path.join(ROOT, '김포한강한의원_업무매뉴얼.xlsx')

FONT = 'Arial'
INK, MUTED = '1C211E', '5B6560'
ACCENT, ACCENT_SOFT = '2E6A4E', 'E9F1EC'
MARK_SOFT = 'F8EEE8'
LINE = 'D8DDD8'

thin = Side(style='thin', color=LINE)
BORDER = Border(bottom=thin)
HEAD_FILL = PatternFill('solid', fgColor=ACCENT)
HEAD_FONT = Font(name=FONT, size=10, bold=True, color='FFFFFF')
BODY = Font(name=FONT, size=10, color=INK)
BODY_MUTED = Font(name=FONT, size=10, color=MUTED)
TOP_WRAP = Alignment(vertical='top', wrap_text=True)

# 구분에 따라 칠하는 색. 급할 때 멘트가 먼저 눈에 들어와야 한다
KIND_FILL = {
    '멘트': PatternFill('solid', fgColor=ACCENT_SOFT),
    '주의': PatternFill('solid', fgColor=MARK_SOFT),
}

def est_height(text, chars_per_line=42):
    lines = 0
    for ln in str(text).split('\n'):
        w = sum(2 if ord(c) > 0x1100 else 1 for c in ln)
        lines += max(1, math.ceil(w / chars_per_line))
    return min(320, max(15, lines * 14 + 4))

def head(ws, cols, widths):
    ws.append(cols)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        c = ws.cell(row=1, column=i)
        c.font, c.fill = HEAD_FONT, HEAD_FILL
        c.alignment = Alignment(vertical='center', horizontal='left')
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = 'A2'

def finish(ws, ncols, content_col):
    last = ws.max_row
    ws.auto_filter.ref = f'A1:{get_column_letter(ncols)}{last}'
    for r in range(2, last + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY if c == content_col else BODY_MUTED
            cell.alignment = TOP_WRAP
            cell.border = BORDER
        kind = ws.cell(row=r, column=3).value if ncols >= 4 else None
        if kind in KIND_FILL:
            for c in range(1, ncols + 1):
                ws.cell(row=r, column=c).fill = KIND_FILL[kind]
        ws.row_dimensions[r].height = est_height(ws.cell(row=r, column=content_col).value or '')

wb = Workbook()

# ── 1. 읽는 법 ───────────────────────────────────────────────
ws = wb.active
ws.title = '읽는 법'
for i, w in enumerate([18, 30, 62], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

def put(r, a, b='', c='', bold=False, size=10, color=INK, fill=None):
    for col, v in ((1, a), (2, b), (3, c)):
        cell = ws.cell(row=r, column=col, value=v)
        cell.font = Font(name=FONT, size=size, bold=bold, color=color)
        cell.alignment = TOP_WRAP
        if fill: cell.fill = PatternFill('solid', fgColor=fill)
    return r + 1

r = 1
r = put(r, '김포한강한의원 업무 매뉴얼', bold=True, size=16, color=ACCENT)
r = put(r, '데스크와 치료실에서 매일 하는 일을 적어놓은 것입니다. 새로 오신 분이 이것만 보고 하루를 돌릴 수 있게 하는 것이 목표입니다.')
r += 1

r = put(r, '이 파일 쓰는 법', bold=True, size=12, fill=ACCENT_SOFT)
for a, b in [
    ('아래 시트 탭', '데스크 · 치료실 · 한약·탕전 · 원무·정산·재고 네 가지가 본문입니다.'),
    ('찾을 때', '맨 윗줄의 필터(▽)를 눌러 「절」이나 「구분」으로 좁히면 됩니다.'),
    ('구분 = 멘트', '환자분께 그대로 하는 말입니다. 초록색으로 칠해 두었습니다.'),
    ('구분 = 주의', '틀리면 문제가 되는 것입니다. 주황색으로 칠해 두었습니다.'),
    ('멘트 모음', '멘트만 한자리에 모아 두었습니다. 전화 받으면서 보기 좋습니다.'),
    ('금액표', '진료비 · 한약 · 비급여 금액을 모아 두었습니다. 금액은 여기서만 고칩니다.'),
    ('확인 필요', '아직 정해지지 않은 것들입니다. 정해지면 적고 「완료」에 √ 하시면 됩니다.'),
    ('그림', '화면 캡처는 이 파일에 넣지 못했습니다. 웹페이지 판에서 보실 수 있습니다.'),
]:
    r = put(r, a, b)
r += 1

r = put(r, '병원 기본 정보', bold=True, size=12, fill=ACCENT_SOFT)
for a, b in [
    ('병원명', '김포한강한의원'), ('대표전화', '031-8049-7541'),
    ('주소', '경기도 김포시 김포한강4로 110 일신프라자 202호'),
    ('오시는 길', '김포골드라인 장기역 3·4번 출구 도보 1분. 다이소 맞은편'),
    ('주차', '건물 옥상주차장, 장기1공영주차장, 장기2공영주차장'),
    ('원장', '이진희 대표원장, 왕소정 대표원장'),
    ('홈페이지', 'https://gimpohangang.com'),
]:
    r = put(r, a, b)
r += 1

r = put(r, '진료시간', bold=True, size=12, fill=ACCENT_SOFT)
for a, b in [
    ('평일', '09:30 – 20:00 (점심 13:00 – 14:00)'),
    ('토요일', '09:30 – 15:00'),
    ('공휴일 · 대체공휴일', '09:30 – 15:00'),
    ('일요일', '휴진'),
    ('설날 당일 · 추석 당일', '휴진'),
]:
    r = put(r, a, b)
r = put(r, '', '공휴일도 진료합니다. 설날과 추석은 당일만 쉬고 연휴 나머지 날은 09:30–15:00 진료합니다.')
r += 1

r = put(r, '모두에게 해당하는 세 가지', bold=True, size=12, fill=ACCENT_SOFT)
for a, b in [
    ('1', '환자분 정보는 병원 밖으로 나가지 않습니다. 대기실에서 이름 외의 정보를 크게 부르지 않습니다.'),
    ('2', '효과를 약속하지 않습니다. 치료 효과와 경과에 대한 설명은 원장님 몫입니다. (의료법 제56조)'),
    ('3', '판단은 원장님이 합니다. 직원이 진단하거나 치료 방향을 답하지 않습니다.'),
]:
    r = put(r, a, b)
r += 1

summary_row = r
r = put(r, '들어 있는 것', bold=True, size=12, fill=ACCENT_SOFT)
count_start = r

# ── 2~5. 본문 네 시트 ────────────────────────────────────────
SHEETS = ['데스크', '치료실', '한약·탕전', '원무·정산·재고']
for name in SHEETS:
    s = wb.create_sheet(name)
    head(s, ['절', '소제목', '구분', '내용'], [22, 24, 8, 100])
    for sec, sub, kind, text in D['docs'][name]:
        s.append([sec, sub or '', kind, text])
    finish(s, 4, 4)

# ── 6. 멘트 모음 ─────────────────────────────────────────────
s = wb.create_sheet('멘트 모음')
head(s, ['문서', '절', '언제', '이렇게 말합니다'], [14, 22, 26, 96])
for name in SHEETS:
    for sec, sub, kind, text in D['docs'][name]:
        if (kind == '멘트' or '멘트' in text
                or text.lstrip().startswith('"')
                or re.search(r'[:：]\s*"', text)):
            body = re.sub(r'^\s*멘트\s*', '', text)
            s.append([name, sec, sub or '', body])
finish(s, 4, 4)
ment_rows = s.max_row - 1

# ── 7. 금액표 ────────────────────────────────────────────────
s = wb.create_sheet('금액표')
head(s, ['문서', '무엇', '항목', '구분', '금액'], [14, 26, 34, 20, 30])
def is_money(t):
    blob = ' '.join(t['head']) + ' ' + ' '.join(' '.join(r) for r in t['body'])
    return ('원' in blob or '%' in blob
            or any(k in (t['sec'] + t['sub']) for k in ('비용', '가격', '금액', '수납', '다이어트 종류')))
for t in D['tables']:
    if not is_money(t):
        continue
    what = t['sub'] or t['sec']
    head_row = t['head']
    for row in t['body']:
        label = row[0]
        for col_name, val in zip(head_row[1:], row[1:]):
            if val:
                s.append([t['doc'], what, label, col_name, val])
# 표 밖에 문장으로 적힌 금액도 빠뜨리지 않는다
for name in SHEETS:
    for sec, sub, kind, text in D['docs'][name]:
        if kind in ('항목', '안내') and re.search(r'\d[\d,]*\s*(원|만원)', text) and '표' not in kind:
            first = text.split('\n')[0]
            s.append([name, sub or sec, first[:120], '', ''])
finish(s, 5, 3)
money_rows = s.max_row - 1

# ── 8. 확인 필요 ─────────────────────────────────────────────
s = wb.create_sheet('확인 필요')
head(s, ['문서', '절', '무엇을 정해야 하나', '정한 내용', '완료'], [14, 24, 78, 44, 8])
STRONG = re.compile(r'확인 필요|정해지지 않|채워\s*주세요|확정해\s*주세요|아직 정해')
SOFT = re.compile(r'확인해서|확인해\s*주세요|확인이 필요|적어\s*주세요|정리해\s*주세요|확인하고 채')
seen = set()
for name in SHEETS:
    for sec, sub, kind, text in D['docs'][name]:
        hit = STRONG.search(text) or (kind == '주의' and SOFT.search(text))
        if hit and (name, text) not in seen:
            seen.add((name, text))
            s.append([name, sub or sec, text, '', ''])
finish(s, 5, 3)
todo_rows = s.max_row - 1
# 노란 칠은 「여기를 채우세요」라는 뜻이다
for r_ in range(2, s.max_row + 1):
    for c_ in (4, 5):
        s.cell(row=r_, column=c_).fill = PatternFill('solid', fgColor='FFF6CC')

# ── 읽는 법의 건수는 수식으로 센다. 내용이 늘면 따라 늘어난다 ──
ws = wb['읽는 법']
r = count_start
rows_map = {n: len(D['docs'][n]) + 1 for n in SHEETS}
for n in SHEETS:
    last = rows_map[n] + 50
    ws.cell(row=r, column=1, value=n).font = Font(name=FONT, size=10, color=MUTED)
    ws.cell(row=r, column=2, value=f"=COUNTA('{n}'!D2:D{last})").font = Font(name=FONT, size=10, color=INK)
    ws.cell(row=r, column=3, value=f"=\"멘트 \"&COUNTIF('{n}'!C2:C{last},\"멘트\")&\"개 · 주의 \"&COUNTIF('{n}'!C2:C{last},\"주의\")&\"개\"").font = Font(name=FONT, size=10, color=MUTED)
    r += 1
for label, sheet, col, tot in [('멘트 모음', '멘트 모음', 'D', ment_rows),
                               ('금액표', '금액표', 'C', money_rows)]:
    ws.cell(row=r, column=1, value=label).font = Font(name=FONT, size=10, color=MUTED)
    ws.cell(row=r, column=2, value=f"=COUNTA('{sheet}'!{col}2:{col}{tot+60})").font = Font(name=FONT, size=10, color=INK)
    r += 1
ws.cell(row=r, column=1, value='확인 필요').font = Font(name=FONT, size=10, color=MUTED)
ws.cell(row=r, column=2, value=f"=COUNTA('확인 필요'!C2:C{todo_rows+60})").font = Font(name=FONT, size=10, color=INK)
ws.cell(row=r, column=3, value=f"=\"아직 정하지 않은 것 \"&(COUNTA('확인 필요'!C2:C{todo_rows+60})-COUNTA('확인 필요'!E2:E{todo_rows+60}))&\"개\"").font = Font(name=FONT, size=10, bold=True, color='9C4A23')
r += 2
ws.cell(row=r, column=1, value='만든 날 2026-09-03 · 원본은 저장소 docs/manual/ 의 마크다운입니다.').font = Font(name=FONT, size=9, color=MUTED)
for rr in range(1, r + 1):
    ws.row_dimensions[rr].height = est_height(ws.cell(row=rr, column=2).value or ws.cell(row=rr, column=1).value or '', 30)

wb.save(OUT)
print('저장:', OUT)
print('멘트', ment_rows, '· 금액', money_rows, '· 확인 필요', todo_rows)
